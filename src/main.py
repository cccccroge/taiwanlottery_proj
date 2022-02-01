import os
from utils.paths import SCRAPY_FOLDER, DATA_FOLDER
from openpyxl import Workbook
from utils.constant import Game
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill
from crawler import Crawler


def main():
    l = crawl_to_list()
    # create_analyzing_excel(l)
    print(l)


def crawl_to_list():
    crawler = Crawler(game_key=Game.BIG_LOTTERY, start_year_month="2022-01")
    crawler.start()
    return crawler.result


def create_analyzing_excel(src_list):
    wb = Workbook()
    construct_sum_sheet(wb, src_list)
    construct_tail_sheet(wb, src_list)
    construct_indi_sheet(wb, src_list)
    wb.save(os.path.join(DATA_FOLDER, f"{Game.GINTSAI_539}.xlsx"))


def construct_sum_sheet(wb, list):
    ws = wb.active
    ws.title = "全部"
    ws.append(get_pricedraw_header())
    set_pricedraw_sheet_style(ws)

    for item in list:
        r = get_row_by_item(item)
        ws.append(r)

    construct_statistic_form(ws)


def construct_tail_sheet(wb, list):
    ws = wb.create_sheet(title="尾數")
    ws.append(get_tail_header())
    ws.append(
        40 * [""]
    )  # set style after will cause no empty line at line 2, so manully add it

    for item in list:
        r = get_tail_row_by_item(item)
        ws.append(r)

    set_tail_sheet_style(ws)
    construct_tail_statistic_form(ws)


def construct_indi_sheet(wb, list):
    ws_indi = []

    for i in range(0, 39):
        ws_indi.append(wb.create_sheet(title=str(i + 1)))
        ws_indi[i].append(get_pricedraw_header())
        set_pricedraw_sheet_style(ws_indi[i])

    for item in list:
        r = get_row_by_item(item)
        for n in item["nums"]:
            ws_indi[int(n) - 1].append(r)

    for i in range(0, 39):
        construct_statistic_form(ws_indi[i])


def get_pricedraw_header():
    return ["日期", *(np.arange(1, 40))]


def get_tail_header():
    list = np.arange(1, 40).tolist()
    list = map(lambda r: tail_rank_to_idx(r), list)
    return ["日期", *(list)]


def set_pricedraw_sheet_style(ws):
    ws.column_dimensions["A"].width = 10
    for i in range(2, 41):
        c = get_column_letter(i)
        ws.column_dimensions[c].width = 4
    ws.freeze_panes = ws["AO2"]


def set_tail_sheet_style(ws):
    ws.column_dimensions["A"].width = 10
    for i in range(2, 41):
        c = get_column_letter(i)
        ws.column_dimensions[c].width = 4
    ws.freeze_panes = ws["AO2"]

    colors = ["EEEEEE", "FFFFFF"]
    j = 0
    for i in range(2, 40, 4):
        cell_range = f"{get_column_letter(i)}1:{get_column_letter(i + 3)}{ws.max_row}"
        set_bg_color(ws, cell_range, colors[j])
        j = (j + 1) % 2
        set_border(ws, cell_range)


def get_row_by_item(item):
    row = [item["date"], *(39 * [""])]
    for n in item["nums"]:
        row[int(n)] = int(n)
    return row


def get_tail_row_by_item(item):
    row = [item["date"], *(39 * [""])]
    for n in item["nums"]:
        n = int(n)
        row[tail_idx_to_rank(n)] = n
    return row


def tail_idx_to_rank(n):
    d1 = n % 10
    d2 = int(n / 10)
    return (4 * (d1 - 1) + d2 + 1) if (n % 10) else (36 + d2)


def tail_rank_to_idx(n):
    d1 = int((n - 1) / 4) + 1
    d2 = (n - 1) % 4
    return d2 * 10 + d1


def construct_statistic_form(ws):
    max_row = ws.max_row
    r = max_row + 5
    num_row = r + 1
    occ_row = r + 2
    rank_row = r + 3

    ws[f"A{num_row}"] = "號碼"
    ws[f"A{occ_row}"] = "次數"
    ws[f"A{rank_row}"] = "排名"
    for i in range(1, 40):
        col = f"{get_column_letter(i + 1)}"
        ws[f"{col}{num_row}"] = i
        formula_count = f"=COUNT({col}3:{col}{max_row})"
        ws[f"{col}{occ_row}"] = formula_count
        formula_rank = f"=RANK({col}{occ_row}, B{occ_row}:AN{occ_row})"
        ws[f"{col}{rank_row}"] = formula_rank

    set_border(ws, f"A{num_row}:AN{rank_row}")
    set_bg_color(ws, f"A{num_row}:A{rank_row}", "DDDDDD")


def construct_tail_statistic_form(ws):
    max_row = ws.max_row
    r = max_row + 5
    num_row = r + 1
    occ_row = r + 2
    rank_row = r + 3

    ws[f"A{num_row}"] = "號碼"
    ws[f"A{occ_row}"] = "次數"
    ws[f"A{rank_row}"] = "排名"
    for i in range(1, 40):
        col = f"{get_column_letter(i + 1)}"
        ws[f"{col}{num_row}"] = tail_rank_to_idx(i)
        formula_count = f"=COUNT({col}3:{col}{max_row})"
        ws[f"{col}{occ_row}"] = formula_count
        formula_rank = f"=RANK({col}{occ_row}, B{occ_row}:AN{occ_row})"
        ws[f"{col}{rank_row}"] = formula_rank

    set_border(ws, f"A{num_row}:AN{rank_row}")
    set_bg_color(ws, f"A{num_row}:A{rank_row}", "DDDDDD")


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def set_bg_color(ws, cell_range, color):
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=color)


if __name__ == "__main__":
    main()
