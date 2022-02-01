import os
from utils.constant import Game, GAME_META
from utils.paths import DATA_FOLDER
from openpyxl import Workbook
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font
from pdb import set_trace


class ExcelExporter:
    def __init__(self, src_list, game_key):
        self.list = src_list
        self.num_max = GAME_META[game_key]["num_max"]
        self.interval = int((self.num_max + 1) / 10)
        self.output_name = game_key.value

    def execute(self):
        self.wb = Workbook()
        self.__construct_sum_sheet()
        self.__construct_tail_sheet()
        self.__construct_indi_sheet()
        self.wb.save(os.path.join(DATA_FOLDER, f"{self.output_name}.xlsx"))

    def __construct_sum_sheet(self):
        ws = self.wb.active
        ws.title = "全部"
        ws.append(self.__get_normal_header())
        self.__set_normal_sheet_style(ws)

        for item in self.list:
            r, special_num_idx = self.__get_row_by_item(item)
            ws.append(r)
            if special_num_idx != None:
                self.__set_emphasized_font(
                    ws, f"{get_column_letter(special_num_idx + 1)}{ws.max_row}"
                )

        self.__construct_statistic_form(ws)

    def __construct_tail_sheet(self):
        ws = self.wb.create_sheet(title="尾數")
        ws.append(self.__get_tail_header())
        ws.append(
            (self.num_max + 1) * [""]
        )  # set style after will cause no empty line at line 2, so manully add it

        for item in self.list:
            r, special_num_idx = self.__get_tail_row_by_item(item)
            ws.append(r)
            if special_num_idx != None:
                self.__set_emphasized_font(
                    ws, f"{get_column_letter(special_num_idx + 1)}{ws.max_row}"
                )

        self.__set_tail_sheet_style(ws)
        self.__construct_tail_statistic_form(ws)

    def __construct_indi_sheet(self):
        ws_indi = []

        for i in range(0, self.num_max):
            ws_indi.append(self.wb.create_sheet(title=str(i + 1)))
            ws_indi[i].append(self.__get_normal_header())
            self.__set_normal_sheet_style(ws_indi[i])

        for item in self.list:
            r, special_num_idx = self.__get_row_by_item(item)
            for n in item["nums"]:
                ws_indi[int(n) - 1].append(r)
            if special_num_idx != None:
                ws = ws_indi[special_num_idx - 1]
                self.__set_emphasized_font(
                    ws, f"{get_column_letter(special_num_idx + 1)}{ws.max_row}"
                )

        for i in range(0, self.num_max):
            self.__construct_statistic_form(ws_indi[i])

    def __get_normal_header(self):
        return ["日期", *(np.arange(1, self.num_max + 1))]

    def __get_tail_header(self):
        list = np.arange(1, self.num_max + 1).tolist()
        list = map(lambda r: self.__tail_idx_to_num(r), list)
        return ["日期", *(list)]

    def __set_normal_sheet_style(self, ws):
        ws.column_dimensions["A"].width = 10
        for i in range(2, self.num_max + 2):
            c = get_column_letter(i)
            ws.column_dimensions[c].width = 4
        ws.freeze_panes = ws[f"B2"]

    def __set_tail_sheet_style(self, ws):
        ws.column_dimensions["A"].width = 10
        for i in range(2, self.num_max + 2):
            c = get_column_letter(i)
            ws.column_dimensions[c].width = 4
        ws.freeze_panes = ws["B2"]

        colors = ["EEEEEE", "FFFFFF"]
        j = 0
        for i in range(2, self.num_max + 1, self.interval):
            cell_range = f"{get_column_letter(i)}1:{get_column_letter(i + self.interval - 1)}{ws.max_row}"
            self.__set_bg_color(ws, cell_range, colors[j])
            j = (j + 1) % 2
            self.__set_border(ws, cell_range)

    def __get_row_by_item(self, item):
        row = [item["date"], *(self.num_max * [""])]
        for n in item["nums"]:
            row[int(n)] = int(n)
        special_num_idx = int(item["special_num"]) if ("special_num" in item) else None

        return row, special_num_idx

    def __get_tail_row_by_item(self, item):
        row = [item["date"], *(self.num_max * [""])]
        for n in item["nums"]:
            n = int(n)
            row[self.__tail_num_to_idx(n)] = n
        special_num_idx = (
            self.__tail_num_to_idx(int(item["special_num"]))
            if ("special_num" in item)
            else None
        )
        return row, special_num_idx

    def __tail_num_to_idx(self, n):
        d1 = n % 10
        d2 = int(n / 10)
        return (
            (self.interval * (d1 - 1) + d2 + 1)
            if (n % 10)
            else (9 * self.interval + d2)
        )

    def __tail_idx_to_num(self, n):
        d1 = int((n - 1) / self.interval) + 1
        d2 = (n - 1) % self.interval
        return d2 * 10 + d1

    def __construct_statistic_form(self, ws):
        max_row = ws.max_row
        r = max_row + 5
        num_row = r + 1
        occ_row = r + 2
        rank_row = r + 3

        ws[f"A{num_row}"] = "號碼"
        ws[f"A{occ_row}"] = "次數"
        ws[f"A{rank_row}"] = "排名"
        for i in range(1, self.num_max + 1):
            col = f"{get_column_letter(i + 1)}"
            ws[f"{col}{num_row}"] = i
            formula_count = f"=COUNT({col}3:{col}{max_row})"
            ws[f"{col}{occ_row}"] = formula_count
            formula_rank = f"=RANK({col}{occ_row}, B{occ_row}:{get_column_letter(self.num_max + 1)}{occ_row})"
            ws[f"{col}{rank_row}"] = formula_rank

        self.__set_border(
            ws, f"A{num_row}:{get_column_letter(self.num_max + 1)}{rank_row}"
        )
        self.__set_bg_color(ws, f"A{num_row}:A{rank_row}", "DDDDDD")

    def __construct_tail_statistic_form(self, ws):
        max_row = ws.max_row
        r = max_row + 5
        num_row = r + 1
        occ_row = r + 2
        rank_row = r + 3

        ws[f"A{num_row}"] = "號碼"
        ws[f"A{occ_row}"] = "次數"
        ws[f"A{rank_row}"] = "排名"
        for i in range(1, self.num_max + 1):
            col = f"{get_column_letter(i + 1)}"
            ws[f"{col}{num_row}"] = self.__tail_idx_to_num(i)
            formula_count = f"=COUNT({col}3:{col}{max_row})"
            ws[f"{col}{occ_row}"] = formula_count
            formula_rank = f"=RANK({col}{occ_row}, B{occ_row}:{get_column_letter(self.num_max + 1)}{occ_row})"
            ws[f"{col}{rank_row}"] = formula_rank

        self.__set_border(
            ws, f"A{num_row}:{get_column_letter(self.num_max + 1)}{rank_row}"
        )
        self.__set_bg_color(ws, f"A{num_row}:A{rank_row}", "DDDDDD")

    def __set_border(self, ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def __set_bg_color(self, ws, cell_range, color):
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=color)

    def __set_emphasized_font(self, ws, cell_range):
        cell = ws[cell_range]
        cell.font = Font(color="FF0000", bold=True)
